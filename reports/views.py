# reports/views.py
import logging
from django.views.generic import TemplateView
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, F
from django.utils import timezone
from datetime import timedelta, datetime
from django.core.exceptions import FieldError
from django.contrib.auth.mixins import LoginRequiredMixin
from io import BytesIO

# For exports
import csv
import json
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

from sales.models import Order, OrderItem
from products.models import Product, Category
from customers.models import Customer

logger = logging.getLogger(__name__)


def get_user_subscription(user):
    """
    Return the ClientSubscription for this user (member or owner),
    or None if none.
    """
    sub = getattr(user, 'subscription', None)
    if sub:
        return sub
    return getattr(user, 'owned_subscription', None)


class ReportsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        sub = get_user_subscription(user)

        # Parse date range / period
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        period = self.request.GET.get('period', 'today')
        today = timezone.now().date()
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = end_date = today
        else:
            if period == 'week':
                start_date = today - timedelta(days=7)
                end_date = today
            elif period == 'month':
                start_date = today - timedelta(days=30)
                end_date = today
            else:  # today or default
                start_date = end_date = today

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
        })

        if not sub:
            # No subscription: zero metrics, empty charts
            context.update({
                'total_sales': 0,
                'total_orders': 0,
                'avg_order_value': 0,
                'total_items_sold': 0,
                'chart_data': [],
                'product_names': [],
                'product_revenues': [],
                'category_names': [],
                'category_sales': [],
            })
            return context

        # Base queryset: completed orders in date range and subscription
        completed_orders = Order.objects.filter(
            status='completed',
            created_at__date__range=[start_date, end_date],
            # If Order has subscription FK:
            subscription=sub
            # Otherwise: filter(user__subscription=sub)
            # Uncomment next line if no subscription FK:
            # user__subscription=sub
        )

        # Sales metrics
        total_sales = completed_orders.aggregate(total=Sum('total'))['total'] or 0
        total_orders = completed_orders.count()

        # Total items sold: sum quantity from OrderItem, scoped by orders in this sub
        order_ids = completed_orders.values_list('id', flat=True)
        total_items_sold = OrderItem.objects.filter(
            order_id__in=order_ids
        ).aggregate(total=Sum('quantity'))['total'] or 0

        avg_order_value = (total_sales / total_orders) if total_orders > 0 else 0

        context.update({
            'total_sales': total_sales,
            'total_orders': total_orders,
            'avg_order_value': avg_order_value,
            'total_items_sold': total_items_sold,
        })

        # Chart data
        chart_data = []
        if period in ('month', 'week', 'today'):
            if period == 'month':
                days = (end_date - start_date).days + 1
                for i in range(days):
                    day = start_date + timedelta(days=i)
                    day_qs = Order.objects.filter(
                        status='completed',
                        created_at__date=day,
                        subscription=sub
                        # or user__subscription=sub
                    )
                    day_total = day_qs.aggregate(total=Sum('total'))['total'] or 0
                    chart_data.append({
                        'date': day.strftime('%m-%d'),
                        'sales': float(day_total)
                    })
            else:
                # week or today: for uniformity, treat today as 1-day week
                num_days = 7 if period == 'week' else 1
                for i in range(num_days):
                    day = start_date + timedelta(days=i)
                    day_qs = Order.objects.filter(
                        status='completed',
                        created_at__date=day,
                        subscription=sub
                    )
                    day_total = day_qs.aggregate(total=Sum('total'))['total'] or 0
                    label = day.strftime('%a') if period == 'week' else day.strftime('%m-%d')
                    chart_data.append({
                        'date': label,
                        'sales': float(day_total)
                    })
        context['chart_data'] = chart_data

        # Top products by revenue in this period & subscription
        # Ensure Product has subscription FK
        top_products_qs = Product.objects.filter(
            subscription=sub,
            orderitem__order__status='completed',
            orderitem__order__created_at__date__range=[start_date, end_date],
            orderitem__order__subscription=sub
        ).annotate(
            revenue=Sum(F('orderitem__quantity') * F('orderitem__price'))
        ).order_by('-revenue')[:5]

        context['product_names'] = [p.name for p in top_products_qs]
        context['product_revenues'] = [float(p.revenue or 0) for p in top_products_qs]

        # Top categories similarly
        top_categories_qs = Category.objects.filter(
            products__subscription=sub,
            products__orderitem__order__status='completed',
            products__orderitem__order__created_at__date__range=[start_date, end_date],
            products__orderitem__order__subscription=sub
        ).annotate(
            total_sales=Sum(F('products__orderitem__quantity') * F('products__orderitem__price'))
        ).order_by('-total_sales')[:5]

        context['category_names'] = [c.name for c in top_categories_qs]
        context['category_sales'] = [float(c.total_sales or 0) for c in top_categories_qs]

        return context


class SalesReportView(LoginRequiredMixin, TemplateView):
    # Renders or exports a detailed sales report scoped by subscription
    template_name = 'reports/dashboard.html'  # fallback if needed

    def get(self, request, *args, **kwargs):
        user = request.user
        sub = get_user_subscription(user)

        # Parse start/end date
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        export_format = request.GET.get('format', 'json')

        # Determine date range
        if start_date_str and end_date_str:
            start_date = end_date = None
            for fmt in ('%Y-%m-%d', '%B %d, %Y', '%m/%d/%Y'):
                try:
                    start_date = datetime.strptime(start_date_str, fmt).date()
                    end_date = datetime.strptime(end_date_str, fmt).date()
                    break
                except Exception:
                    continue
            if not (start_date and end_date):
                end_date = timezone.now().date()
                start_date = end_date - timedelta(days=30)
        else:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)

        if not sub:
            # No subscription: return empty or error
            empty_data = {
                'period': f"{start_date} to {end_date}",
                'total_sales': 0,
                'total_orders': 0,
                'total_items': 0,
                'orders': []
            }
            if export_format == 'csv':
                return self.export_csv(empty_data)
            elif export_format == 'pdf':
                return self.export_pdf(empty_data)
            elif export_format == 'excel':
                return self.export_excel(empty_data)
            else:
                return JsonResponse(empty_data)

        # Filter orders by subscription & date & completed
        orders_qs = Order.objects.filter(
            status='completed',
            created_at__date__range=(start_date, end_date),
            subscription=sub
        ).prefetch_related('items__product')

        total_sales = orders_qs.aggregate(total=Sum('total'))['total'] or 0
        total_orders = orders_qs.count()
        total_items = orders_qs.aggregate(total_items=Sum('items__quantity'))['total_items'] or 0

        report_data = {
            'period': f"{start_date} to {end_date}",
            'total_sales': float(total_sales),
            'total_orders': total_orders,
            'total_items': total_items,
            'orders': []
        }

        for order in orders_qs:
            items_count = order.items.aggregate(total=Sum('quantity'))['total'] or 0
            customer_name = order.customer.name if order.customer else 'Walk-in'
            report_data['orders'].append({
                'order_id': order.id,
                'date': order.created_at.strftime('%Y-%m-%d'),
                'total': float(order.total),
                'items_count': items_count,
                'customer': customer_name,
            })

        # Export
        if export_format == 'csv':
            return self.export_csv(report_data)
        elif export_format == 'pdf':
            return self.export_pdf(report_data)
        elif export_format == 'excel':
            return self.export_excel(report_data)
        else:
            return JsonResponse(report_data)

    def export_csv(self, data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="sales_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Sales Report Summary'])
        writer.writerow(['Period', data['period']])
        writer.writerow(['Total Sales', f"${data['total_sales']:.2f}"])
        writer.writerow(['Total Orders', data['total_orders']])
        writer.writerow(['Total Items', data['total_items']])
        writer.writerow([])
        writer.writerow(['Order ID', 'Date', 'Total', 'Items Count', 'Customer'])
        for order in data['orders']:
            writer.writerow([
                order['order_id'],
                order['date'],
                f"${order['total']:.2f}",
                order['items_count'],
                order['customer']
            ])
        return response

    def export_pdf(self, data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, f"Sales Report: {data['period']}")

        # Summary
        p.setFont("Helvetica", 12)
        y = height - 100
        p.drawString(50, y, f"Total Sales: ${data['total_sales']:,.2f}")
        p.drawString(50, y - 20, f"Total Orders: {data['total_orders']:,}")
        p.drawString(50, y - 40, f"Total Items Sold: {data['total_items']:,}")
        avg_val = data['total_sales'] / max(data['total_orders'], 1)
        p.drawString(50, y - 60, f"Average Order Value: ${avg_val:,.2f}")

        # Table header
        y -= 100
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Order ID")
        p.drawString(120, y, "Date")
        p.drawString(200, y, "Total")
        p.drawString(280, y, "Items")
        p.drawString(340, y, "Customer")

        # Table rows (limit to first 20)
        p.setFont("Helvetica", 9)
        for order in data['orders'][:20]:
            y -= 15
            if y < 50:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica-Bold", 10)
                p.drawString(50, y, "Order ID")
                p.drawString(120, y, "Date")
                p.drawString(200, y, "Total")
                p.drawString(280, y, "Items")
                p.drawString(340, y, "Customer")
                p.setFont("Helvetica", 9)
                y -= 15
            p.drawString(50, y, str(order['order_id']))
            p.drawString(120, y, order['date'])
            p.drawString(200, y, f"${order['total']:,.2f}")
            p.drawString(280, y, str(order['items_count']))
            p.drawString(340, y, order['customer'])

        p.save()
        response.write(buffer.getvalue())
        buffer.close()
        return response

    def export_excel(self, data):
        if not openpyxl:
            return self.export_csv(data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Summary
        ws['A1'] = 'Sales Report Summary'
        ws['A2'] = 'Period'
        ws['B2'] = data['period']
        ws['A3'] = 'Total Sales'
        ws['B3'] = f"${data['total_sales']:.2f}"
        ws['A4'] = 'Total Orders'
        ws['B4'] = data['total_orders']
        ws['A5'] = 'Total Items'
        ws['B5'] = data['total_items']

        # Header
        ws.append([])
        ws.append(['Order ID', 'Date', 'Total', 'Items Count', 'Customer'])
        for order in data['orders']:
            ws.append([
                order['order_id'],
                order['date'],
                f"${order['total']:.2f}",
                order['items_count'],
                order['customer']
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
        wb.save(response)
        return response

    